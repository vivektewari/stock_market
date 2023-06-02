
import glob
import sys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import pandas as pd
import warnings
from bs4 import BeautifulSoup
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import time
import warnings
import os
from multiprocessing import Pool, Process, cpu_count, Manager
from sql_update import sql_postman
start=time.time()

sql_postman_ = sql_postman(host="localhost", user="vivek", password="password", database="mydb",
                               conversion_dict='/home/pooja/PycharmProjects/stock_valuation/codes/sql_update/codes_to_sql_dict.csv')
dc=sql_postman_.sql_dict
myresult = sql_postman_.read("""select {} from stock_space""".format(sql_postman_.sql_dict['nse_id']))
nse_ids = list(set([myresult[i][0] for i in range(len(myresult))]))
def check_exists_by_xpath(driver,xpath):
    try:
         q=driver.find_element(By.CSS_SELECTOR, xpath)
         return True,q
    except NoSuchElementException:
        return False,-1

def scrape_table(page_source, stock_name, sheet_name,path=None,next_page=False):
    """ Scrape a table from a web page """

    soup = BeautifulSoup(page_source,"lxml")
    table = soup.find('table', {'class': 'tblchart'})
    try:rows = table.find_all('tr')
    except:
        warnings.warn('{} doesnot have {}'.format(stock_name,sheet_name))
        return -1
    if next_page:
        wb = openpyxl.load_workbook(path+'.xlsx')#.format(stock_name)

    else:
        wb = openpyxl.Workbook()


    sheet = wb.active
    sheet.title = sheet_name
    sheet = wb.active
    if next_page:
        first_empty_row = sheet.max_row-2
        for i, col in enumerate(rows):
            if i > 1:
                for j, el in enumerate(col):

                    cell_ref = sheet.cell(first_empty_row+i + 1, j + 1)
                    cell_ref.value = el.string
    else:
        for row in rows:
                row_list = [el.string for el in row]
                sheet.append(row_list)

    wb.save(path + '.xlsx')  # .format(stock_name)

def parse_url(file, i, save_path):
    #driver = webdriver.Chrome("/usr/lib/chromium-browser/chromedriver", options=options)

    nse_id = file.loc[i, 'nse_id']
    spath = save_path + '/' + nse_id + "/"
    if not os.path.exists(spath):
        #uncomment below line if you think some files have been missed
        #spath = save_path + '/faults/' + nse_id + "/" #adjustement
        os.mkdir(spath)



    else:
        return
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    url = file.loc[i, 'price_link']
    try:
        driver.get(url)
        #time.sleep(5)
    except:
        warnings.warn('{} couldnt get th url'.format(nse_id))
        return
    Select(driver.find_element(By.NAME, "ex")).select_by_index(1)
    Select(driver.find_element(By.NAME, "frm_dy")).select_by_index(0)
    Select(driver.find_element(By.NAME, "frm_mth")).select_by_index(0)
    Select(driver.find_element(By.NAME, "frm_yr")).select_by_index(2)
    Select(driver.find_element(By.NAME, "to_dy")).select_by_index(30)
    Select(driver.find_element(By.NAME, "to_mth")).select_by_index(11)
    Select(driver.find_element(By.NAME, "to_yr")).select_by_index(1)
    driver.find_element(By.CSS_SELECTOR, "input[src*=go_btn]").click()
    # time.sleep(5)
    page_exist=True
    next_page=False
    while page_exist:

        page_source = driver.page_source
        scrape_table(page_source, nse_id, 'price', path=spath + 'price', next_page=next_page)
        page_exist,element=check_exists_by_xpath(driver,'a[class="nextprev"]')

        if page_exist: # if next page exist got to next page
            driver.get(element.get_attribute("href")) #element.click()
            next_page=True

    driver.quit()

options = Options()
options.add_argument('--disable-notifications')
options.add_argument("--headless");



part=2
start=time.time()
if part == 1:
    file = '/home/pooja/PycharmProjects/stock_valuation/data/raw_data/money_control_link_file/all_links.csv'
    file=pd.read_csv(file)
    save_path='/home/pooja/PycharmProjects/stock_valuation/data/raw_data/money_control/price_links.csv'
    #pd.DataFrame(columns=['nse_id','price_link']).to_csv(save_path,index=False)
    nse_done=list(pd.read_csv(save_path)['nse_id'])

    driver = webdriver.Chrome("/usr/lib/chromium-browser/chromedriver", options=options)
    for i in range(len(file)):
        dict = {}
        dict['nse_id'] = [file.loc[i, 'nse_id']]
        if dict['nse_id'][0] in nse_done:continue
        url = file.loc[i, 'link']
        try:driver.get(url)
        except:
            warnings.warn('{} couldnt get th url'.format(dict['nse_id']))
            continue
        time.sleep(5)
        try:driver.find_element(By.ID,"priceVolLink").click()
        except:
            warnings.warn('{} couldnt get th historical data url'.format(dict['nse_id']))
            continue
        #time.sleep(5)
        driver.switch_to.window(driver.window_handles[1])
        temp=driver.current_url
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

        if temp != temp:
            warnings.warn('{} no data link found'.format(dict['nse_id']))
            continue
        else:
            dict['price_link'] = [temp]
            pd.DataFrame(dict).to_csv(save_path,mode='a',header=False,index=False)
        print(i,time.time()-start)



if part==2:
    file = '/home/pooja/PycharmProjects/stock_valuation/data/raw_data/money_control_link_file/price_links.csv'
    file = pd.read_csv(file)
    #update
    save_path = '/home/pooja/PycharmProjects/stock_valuation/data/raw_data/web_scrapped/money_control/price/01_jan_2021__31_12_2022/'

    cores = cpu_count()
    pool = Pool(processes=cores)
    batch = 8
    for i in range(len(file)):
        #parse_url(file, i, save_path)
        nse_id=file.loc[i, 'nse_id']
        if nse_id not in nse_ids:continue
        #parse_url(file, i, save_path)
        pool.apply_async(parse_url, args=(file, i, save_path))
        if i % int(
                batch) == 0:  # used so that limited process run in memmory. So batch should be cosen considering availibilty of memmory
            pool.close()
            pool.join()
            pool = Pool(processes=cores)
            print(i, time.time() - start)
    pool.close()
    pool.join()

print("time taken in seconds:{}".format(time.time()-start))