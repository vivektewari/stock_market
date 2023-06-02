import pandas as pd
import requests
import numpy as np
from bs4 import BeautifulSoup
import os
import time
import csv
import glob
import openpyxl
from multiprocessing import Pool, Process, cpu_count, Manager
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
import warnings
from selenium.common.exceptions import NoSuchElementException
def get_html(url):
    """ Get the HTML of a URL """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml') # Use lxml parser
    return soup
start=time.time()
soup = get_html('https://www.moneycontrol.com/stocks/marketinfo/splits/index.php')
time.sleep(5)
options = Options()
options.add_argument('--disable-notifications')
options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
def parse_nse_id(row):
    try:
        soup = get_html('http://moneycontrol.com' + row.find('a', {'class': 'bl_12'}).get('href'))
        spans = soup.find('ul', {'class': 'comp_inf company_slider'}).find_all('span')
    except:
        warnings.warn('not found this link {}'.format(row.find('a', {'class': 'bl_12'}).get('href')))
        return None

    nse_id = np.NAN
    for c in spans:
        if c.text == 'NSE:':
            temp = c.parent.find('p').text
            if len(temp) > 0:
                nse_id = temp
                # print(temp)
            break
    if nse_id==nse_id:return nse_id
    else:return None
def scrape_table(page_source,next_page=False,path=None):
    """ Scrape a table from a web page """

    soup = BeautifulSoup(page_source,"lxml")
    table = soup.find('table', {'class': 'b_12 dvdtbl tbldata14'})
    try:rows = table.find_all('tr')
    except:
        warnings.warn('doesnot have ')
        return -1

    if next_page:
        wb = openpyxl.load_workbook(path+'.xlsx')#.format(stock_name)
    else:
        wb = openpyxl.Workbook()


    sheet = wb.active
    loop=0
    first_empty_row=0
    loop=0
    for i, row in enumerate(rows):

            if i==0:
                nse_id='nse_id'
                if next_page:
                    first_empty_row = sheet.max_row - 2
                    continue
            else:nse_id=parse_nse_id(row)
            if nse_id is not None:
                for j, el in enumerate(row):
                    cell_ref = sheet.cell(first_empty_row + loop + 1, j + 1)
                    if j==0:cell_ref.value=nse_id
                    else:cell_ref.value = el.string
                loop+=1



            #     row_list = [el.text for el in row][2:]
            #     row_list=[nse_id]+row_list
            #     sheet.append(row_list)
            # else:
            #     continue


    wb.save(path + '.xlsx')  # .format(stock_name)


url = 'https://www.moneycontrol.com/stocks/marketinfo/splits/index.php'

try:
        driver.get(url)
        #time.sleep(5)
except:
        warnings.warn('{} couldnt get th url'.format(nse_id))
next_page=False
for i in range(23,24):
    Select(driver.find_element(By.NAME, "sel_year")).select_by_index(i)
    #driver.find_element(By.ID, "bonus_frm").click()
    driver.find_element(By.CSS_SELECTOR, "input[src*=gobtn]").click()
    page_source = driver.page_source
    scrape_table(page_source, next_page, path='/home/pooja/PycharmProjects/stock_valuation/data/raw_data/web_scrapped/money_control/stock_splits/splits')
    next_page=True


if __name__ == '__main__':
   # main()
    print("time taken in seconds:{}".format(time.time() - start))