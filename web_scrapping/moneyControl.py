import warnings

import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup
import os
import time
import csv
import glob
import openpyxl
from multiprocessing import Pool, Process, cpu_count, Manager
def get_html(url):
    """ Get the HTML of a URL """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml') # Use lxml parser
    return soup



def print_csv_columns():
    """ Print the contents of all CSVs """
    for filename in glob.glob('*.csv'):
        with open(filename, newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            for row in reader:
                if row[0]:
                    print(row[0], row[1])
                else:
                    continue

def scrape_table(url, stock_name, sheet_name, next_page = False,path=None):
    """ Scrape a table from a web page """
    soup = get_html(url)
    table = soup.find('table', {'class': 'mctable1'})
    try:rows = table.find_all('tr')
    except:
        warnings.warn('{} doesnot have {}'.format(stock_name,sheet_name))
        return -1

    if next_page:
        wb = openpyxl.load_workbook(path+'.xlsx')#.format(stock_name)

    else:
        wb = openpyxl.Workbook()

    sheet = wb.active
    sheet.title = sheet_name
    sheet = wb.active
    if next_page:
        first_empty_col = sheet.max_column - 2
        for i, row in enumerate(rows):
            for j, el in enumerate(row):
                if j > 2 and j < len(row) - 3:
                    cell_ref = sheet.cell(i + 1, first_empty_col + j + 1)
                    cell_ref.value = el.string
    else:
        for row in rows:
            row_list = [el.string for el in row][:-2]
            sheet.append(row_list)


    wb.save(path+'.xlsx')#.format(stock_name)
    d=0


def create_csv(save_path,dict):
    """ Create a CSV file for each stock price letter """


    dict = {e: '' for e in dict}
    csvfile =open(save_path , 'w', newline='')
    writer = csv.writer(csvfile)
    writer.writerow(dict.keys())
    for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        try:
            soup = get_html('https://www.moneycontrol.com/india/stockpricequote/' + letter)
            time.sleep(5)

            links = soup.find_all('a', {'class': 'bl_12'})



            for link in links:
                dict['text']=link.text
                dict['link']=link['href']
                d=scrape_quick_links(link['href'])
                if d  is None:continue
                for key in d.keys():
                    if key in dict.keys(): dict[key]=d[key]
                writer.writerow(list(dict.values()))

            print("Success for ", letter)

        except Exception as e:
            print("Exception for ", letter, ": ", e)
    csvfile.close()
def print_csv_columns():
    """ Print the contents of all CSVs """
    for filename in glob.glob('*.csv'):
        with open(filename, newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            for row in reader:
                if row[0]:
                    print(row[0], row[1])
                else:
                    continue

def get_html(url):
    """ Get the HTML of a URL """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml') # Use lxml parser
    return soup

def scrape_quick_links(url):
    """ Scrape quick links from a web page """

    soup = get_html(url)
    spans = soup.find('ul',{'class':'comp_inf company_slider'}).find_all('span')
    nse_id=np.NAN
    for c in spans:
        if c.text=='NSE:':
            temp=c.parent.find('p').text
            if len(temp)>0:
                nse_id=temp
                #print(temp)
            break


    if nse_id!=nse_id :
        #print('got nan')
        return None

    quick_links = soup.find('div', {'class': 'quick_links clearfix'})

    links = quick_links.find_all('a')
    dict={}
    dict['nse_id']=nse_id
    for link in links:
        dict[link.text]= link['href']

    #print(dict)
    return dict
def get_active_href(url):
    """ Get the URL of the active page """
    soup = get_html(url)
    try:
        span_tag = soup.find('span', {'class': 'nextpaging'})
        parent_tag = span_tag.find_previous('a')
    except:
        return

    if parent_tag:
        href = parent_tag.get('href')
        if href and href != 'javascript:void();':
            return href
    return None
def parse_url(url,):
    while url:
        # print(url)
        if first_entry:
            r = scrape_table(url, stock_name, sheet_name, path=spath + col)
            if r == -1: r = scrape_table(old_url, stock_name, sheet_name, path=spath + col)
            if r == -1: break
            first_entry = False
        else:
            scrape_table(url, stock_name, sheet_name, True, path=spath + col)
        # print(url)
        url = get_active_href(url)
def main():
    part=2
    """ Get all the stock web links from A-Z available on
    moneycontrol and store them into
    csv files alphabetically """
    column_list = ['text', 'link', 'nse_id', 'Balance Sheet', 'Profit & Loss', 'Quarterly Results',
                   'Half Yearly Results',
                   'Nine Months Results', 'Yearly Results', 'Cash Flows', 'Ratios', 'Capital Structure']
    if part==1:
        file='/home/pooja/PycharmProjects/stock_valuation/data/raw_data/money_control/all_links.csv'
        create_csv(file,dict=column_list[:])


    if part==2:
        """ Read a url for stock and scrape the urls of financial section
        like Balance Sheet, Profit and Loss and, Quarterly an Yearly results,
        Cashflow statments, etc. """


        """ Ge the financial section data which we want. In this case Balance Sheet of the stock """
        file = '/home/pooja/PycharmProjects/stock_valuation/data/raw_data/money_control/all_links_.csv'
        file=pd.read_csv(file)
        parent_folder='/home/pooja/PycharmProjects/stock_valuation/data/to_sql/finacials/'
        cores = cpu_count()
        pool = Pool(processes=cores)
        for i in range(len(file)):
            folder_name=file.loc[i,'nse_id']
            spath = parent_folder+'/'+folder_name+'/'
            if not os.path.exists(spath):os.mkdir(spath)
            else:continue
            print(folder_name)
            for col in column_list:
                if col in ['text', 'link', 'nse_id'] :continue
                url=file.loc[i,col]
                old_url=url
                url_array=url.split("/")
                url_array[-2]='consolidated-'+url_array[-2]
                url=""
                for el in url_array:
                    url=url+el+"/"


                stock_name=folder_name
                sheet_name=col
                first_entry = True


                    #print(url)




    # url = "https://www.moneycontrol.com/financials/relianceindustries/balance-sheetVI/RI#RI"
    # stock_name = "RELIANCE"
    # sheet_name = "Balance Sheet"
    # #
    # """ Store all the available data of all years including searching for previous years's data if available
    # and saving them into a excel file with sheet Balance Sheet in this case """
    # spath='/home/pooja/PycharmProjects/stock_valuation/data/raw_data/money_control/'
    # #scrape_table(url, stock_name, sheet_name,)



if __name__ == '__main__':
    main()